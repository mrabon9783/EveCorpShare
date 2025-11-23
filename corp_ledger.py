#!/usr/bin/env python3
import argparse
import json
import sqlite3
import time
from dataclasses import dataclass
from typing import Any, Dict, Optional, List, Tuple
import openpyxl
import os
from datetime import datetime

import requests
import yaml

# -----------------------
# Share rules (adjust here)
# -----------------------
# 1,000,000,000 ISK => 1 share
SHARE_UNIT_ISK = 1_000_000_000

# -----------------------
# Config + ESI client
# -----------------------

@dataclass
class ESIConfig:
    client_id: str
    client_secret: str
    refresh_token: str
    base_url: str
    token_url: str
    corporation_id: int
    wallet_division: int
    db_path: str = "./ledger.db"
    # Janice integration (optional)
    janice_api_key: Optional[str] = None
    janice_url: str = "https://janice.e-351.com/api/rest/v1/pricer"

    # New fields here:
    custom_discord_webhook: Optional[str] = None
    custom_share_unit_isk: int = 1_000_000_000
    custom_alert_threshold: int = 50_000_000

class ESIClient:
    def __init__(self, cfg: ESIConfig):
        self.cfg = cfg
        self._access_token: Optional[str] = None
        self._token_expiry: float = 0.0

    def _refresh_access_token(self) -> None:
        data = {
            "grant_type": "refresh_token",
            "refresh_token": self.cfg.refresh_token,
        }
        auth = (self.cfg.client_id, self.cfg.client_secret)
        resp = requests.post(self.cfg.token_url, data=data, auth=auth, timeout=30)
        resp.raise_for_status()
        payload = resp.json()
        self._access_token = payload["access_token"]
        self._token_expiry = time.time() + payload.get("expires_in", 1200) - 60

    def _get_access_token(self) -> str:
        if not self._access_token or time.time() >= self._token_expiry:
            self._refresh_access_token()
        return self._access_token

    def get(self, path: str, params: Dict[str, Any] = None) -> Any:
        if params is None:
            params = {}
        headers = {
            "Authorization": f"Bearer {self._get_access_token()}",
            "Accept": "application/json",
        }
        url = f"{self.cfg.base_url}{path}"
        resp = requests.get(url, headers=headers, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json(), resp.headers


def SayItInDiscord(comment: str, cfg: ESIConfig)->None:
    webhook_url = cfg.custom_discord_webhook

    data = {
        "content": f"Update from MB Banking Ledger Bot! 🚀 \r\n{comment}",
        "username": "Ledger Bot"
    }

    requests.post(webhook_url, json=data)
# -----------------------
# DB helpers
# -----------------------
def export_all_to_excel(conn: sqlite3.Connection, output_path: str) -> None:
    """
    Export all key tables into a single Excel workbook with multiple tabs.
    """
    tables = [
        "wallet_journal",
        "donations",
        "contracts",
        "industry_jobs",
        "market_orders",
        "contract_items",
        "type_names",
        "member_flows",
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cur = conn.cursor()

    for t in tables:
        ws = wb.create_sheet(title=t)

        # Get column names
        cur.execute(f"PRAGMA table_info({t})")
        cols = [row[1] for row in cur.fetchall()]
        for c_idx, colname in enumerate(cols, start=1):
            ws.cell(row=1, column=c_idx, value=colname)

        # Write rows
        cur.execute(f"SELECT * FROM {t}")
        rows = cur.fetchall()
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        print(f"  Exported {len(rows)} rows from {t}")

    wb.save(output_path)
    print(f"\nExcel workbook saved to: {output_path}")

def _ensure_contracts_janice_columns(conn: sqlite3.Connection) -> None:
    """
    Ensure contracts table has janice_immediate_split.
    Safe for existing DBs.
    """
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(contracts)")
    cols = [row[1] for row in cur.fetchall()]

    if "janice_immediate_split" not in cols:
        cur.execute("ALTER TABLE contracts ADD COLUMN janice_immediate_split REAL")

    conn.commit()


def init_db(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)

    # Wallet journal
    conn.execute("""
        CREATE TABLE IF NOT EXISTS wallet_journal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            esi_id INTEGER UNIQUE,
            date TEXT,
            ref_type TEXT,
            amount REAL,
            balance REAL,
            description TEXT,
            first_party_id INTEGER,
            second_party_id INTEGER,
            division INTEGER,
            raw_json TEXT
        )
    """)

    # Donations
    conn.execute("""
        CREATE TABLE IF NOT EXISTS donations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            journal_esi_id INTEGER UNIQUE,
            character_id INTEGER,
            amount REAL,
            description TEXT,
            processed INTEGER DEFAULT 0,
            notes TEXT
        )
    """)

    # Contracts
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER UNIQUE,
            issuer_id INTEGER,
            issuer_corporation_id INTEGER,
            assignee_id INTEGER,
            acceptor_id INTEGER,
            start_location_id INTEGER,
            end_location_id INTEGER,
            type TEXT,
            status TEXT,
            title TEXT,
            for_corporation INTEGER,
            date_issued TEXT,
            date_expired TEXT,
            date_accepted TEXT,
            days_to_complete INTEGER,
            price REAL,
            reward REAL,
            collateral REAL,
            volume REAL,
            raw_json TEXT
        )
    """)
    # Make sure Janice columns exist (handles upgrades)
    _ensure_contracts_janice_columns(conn)

    # Contract items
    conn.execute("""
        CREATE TABLE IF NOT EXISTS contract_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER,
            record_id INTEGER,
            type_id INTEGER,
            quantity INTEGER,
            quantity_remaining INTEGER,
            is_included INTEGER,
            is_singleton INTEGER,
            raw_json TEXT,
            UNIQUE(contract_id, record_id)
        )
    """)
    # Character name cache
    conn.execute("""
        CREATE TABLE IF NOT EXISTS characters (
            character_id INTEGER PRIMARY KEY,
            name TEXT,
            last_updated TEXT
        )
    """)

    # Type names cache (for printing and Janice lines)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS type_names (
            type_id INTEGER PRIMARY KEY,
            name TEXT
        )
    """)

    # Member flows (value in/out per character)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS member_flows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            character_id INTEGER NOT NULL,
            direction TEXT NOT NULL,          -- 'in' or 'out'
            source TEXT NOT NULL,             -- 'wallet', 'contract_in', 'contract_out', etc.
            contract_id INTEGER,
            journal_esi_id INTEGER,
            value_isk REAL NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            note TEXT
        )
    """)

    # Industry jobs (corp manufacturing)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS industry_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER UNIQUE,
            installer_id INTEGER,
            facility_id INTEGER,
            activity_id INTEGER,
            blueprint_id INTEGER,
            blueprint_type_id INTEGER,
            product_type_id INTEGER,
            runs INTEGER,
            cost REAL,
            status TEXT,
            start_date TEXT,
            end_date TEXT,
            pause_date TEXT,
            completed_character_id INTEGER,
            completed_date TEXT,
            successful_runs INTEGER,
            location_id INTEGER,
            raw_json TEXT
        )
    """)
    # Market orders (corp orders, open + history)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS market_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER UNIQUE,
            type_id INTEGER,
            location_id INTEGER,
            volume_total INTEGER,
            volume_remain INTEGER,
            price REAL,
            is_buy_order INTEGER,
            issued_by INTEGER,
            state TEXT,
            issued TEXT,
            last_updated TEXT,
            range TEXT,
            wallet_division INTEGER,
            region_id INTEGER,
            is_history INTEGER,   -- 0 = current, 1 = from history endpoint
            raw_json TEXT
        )
    """)

    conn.commit()
    return conn


# -----------------------
# Type name resolution
# -----------------------

def get_type_name(conn: sqlite3.Connection, esi: ESIClient, type_id: int) -> str:
    """
    Resolve an item type_id to its name using ESI /universe/types/{type_id},
    cached in SQLite so we don't spam ESI.
    """
    cur = conn.cursor()
    cur.execute("SELECT name FROM type_names WHERE type_id = ?", (type_id,))
    row = cur.fetchone()
    if row and row[0]:
        return row[0]

    # Not in cache -> fetch from ESI
    path = f"/universe/types/{type_id}/"
    try:
        data, _ = esi.get(path)
    except requests.HTTPError as ex:
        print(f"Failed to resolve type_id {type_id}: {ex}")
        name = f"type:{type_id}"
    else:
        name = data.get("name") or f"type:{type_id}"

    cur.execute(
        "INSERT OR REPLACE INTO type_names (type_id, name) VALUES (?, ?)",
        (type_id, name),
    )
    conn.commit()
    return name

def get_character_name(conn: sqlite3.Connection, esi: ESIClient, character_id: int) -> str:
    """
    Resolve a character_id to its name using ESI /characters/{character_id}/,
    cached in SQLite so we don't spam ESI.
    """
    if character_id is None:
        return "Unknown"

    cur = conn.cursor()
    cur.execute("SELECT name FROM characters WHERE character_id = ?", (character_id,))
    row = cur.fetchone()
    if row and row[0]:
        return row[0]

    # Not in cache -> fetch from ESI
    path = f"/characters/{character_id}/"
    try:
        data, _ = esi.get(path)
        name = data.get("name") or f"char:{character_id}"
    except Exception as ex:
        print(f"Failed to resolve character_id {character_id}: {ex}")
        name = f"char:{character_id}"

    cur.execute("""
        INSERT OR REPLACE INTO characters (character_id, name, last_updated)
        VALUES (?, ?, ?)
    """, (character_id, name, datetime.utcnow().isoformat(timespec="seconds")))
    conn.commit()
    return name
# -----------------------
# Wallet journal + donations
# -----------------------

def sync_wallet_journal(conn: sqlite3.Connection, esi: ESIClient) -> None:
    cur = conn.cursor()
    corp_id = esi.cfg.corporation_id
    division = esi.cfg.wallet_division

    page = 1
    new_count = 0

    while True:
        path = f"/corporations/{corp_id}/wallets/{division}/journal/"
        data, headers = esi.get(path, params={"page": page})
        if not data:
            break

        for row in data:
            esi_id = row["id"]
            cur.execute(
                "SELECT 1 FROM wallet_journal WHERE esi_id = ?",
                (esi_id,),
            )
            if cur.fetchone():
                continue

            cur.execute("""
                INSERT INTO wallet_journal (
                    esi_id, date, ref_type, amount, balance,
                    description, first_party_id, second_party_id,
                    division, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                esi_id,
                row.get("date"),
                row.get("ref_type"),
                row.get("amount"),
                row.get("balance"),
                row.get("reason") or row.get("description"),
                row.get("first_party_id"),
                row.get("second_party_id"),
                division,
                json.dumps(row),
            ))
            new_count += 1

        x_pages = headers.get("X-Pages")
        if x_pages is None or page >= int(x_pages):
            break
        page += 1

    conn.commit()
    print(f"Synced wallet journal: {new_count} new entries.")


def update_donations_from_journal(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute("""
        SELECT esi_id, raw_json
        FROM wallet_journal
        WHERE ref_type = 'player_donation'
          AND esi_id NOT IN (SELECT journal_esi_id FROM donations)
    """)
    rows = cur.fetchall()

    for esi_id, raw_json in rows:
        data = json.loads(raw_json)
        character_id = data.get("first_party_id")
        amount = data.get("amount")
        desc = data.get("reason") or data.get("description") or ""

        cur.execute("""
            INSERT INTO donations (journal_esi_id, character_id, amount, description)
            VALUES (?, ?, ?, ?)
        """, (esi_id, character_id, amount, desc))

    conn.commit()
    print(f"Derived {len(rows)} new donations from wallet journal.")


# -----------------------
# Contracts + items
# -----------------------

def sync_contracts(conn: sqlite3.Connection, esi: ESIClient) -> None:
    cur = conn.cursor()
    corp_id = esi.cfg.corporation_id

    page = 1
    new_contracts = 0

    while True:
        path = f"/corporations/{corp_id}/contracts/"
        data, headers = esi.get(path, params={"page": page})
        if not data:
            break

        for row in data:
            contract_id = row["contract_id"]
            cur.execute("SELECT 1 FROM contracts WHERE contract_id = ?", (contract_id,))
            exists = cur.fetchone()

            if not exists:
                cur.execute("""
                    INSERT INTO contracts (
                        contract_id,
                        issuer_id,
                        issuer_corporation_id,
                        assignee_id,
                        acceptor_id,
                        start_location_id,
                        end_location_id,
                        type,
                        status,
                        title,
                        for_corporation,
                        date_issued,
                        date_expired,
                        date_accepted,
                        days_to_complete,
                        price,
                        reward,
                        collateral,
                        volume,
                        raw_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    contract_id,
                    row.get("issuer_id"),
                    row.get("issuer_corporation_id"),
                    row.get("assignee_id"),
                    row.get("acceptor_id"),
                    row.get("start_location_id"),
                    row.get("end_location_id"),
                    row.get("type"),
                    row.get("status"),
                    row.get("title"),
                    int(row.get("for_corporation") or 0),
                    row.get("date_issued"),
                    row.get("date_expired"),
                    row.get("date_accepted"),
                    row.get("days_to_complete"),
                    row.get("price"),
                    row.get("reward"),
                    row.get("collateral"),
                    row.get("volume"),
                    json.dumps(row),
                ))
                new_contracts += 1

            # Sync items for each contract
            sync_contract_items_for_contract(conn, esi, contract_id)

        x_pages = headers.get("X-Pages")
        if x_pages is None or page >= int(x_pages):
            break
        page += 1

    conn.commit()
    print(f"Synced contracts: {new_contracts} new contracts.")


def sync_contract_items_for_contract(conn: sqlite3.Connection, esi: ESIClient, contract_id: int) -> None:
    cur = conn.cursor()
    corp_id = esi.cfg.corporation_id
    path = f"/corporations/{corp_id}/contracts/{contract_id}/items/"

    try:
        data, headers = esi.get(path)
    except requests.HTTPError as ex:
        print(f"Failed to get items for contract {contract_id}: {ex}")
        return

    new_items = 0
    for item in data:
        record_id = item.get("record_id")
        type_id = item.get("type_id")
        quantity = item.get("quantity", 0)
        quantity_remaining = item.get("quantity_remaining", 0)
        is_included = 1 if item.get("is_included") else 0
        is_singleton = 1 if item.get("is_singleton") else 0

        cur.execute("""
            INSERT OR IGNORE INTO contract_items (
                contract_id,
                record_id,
                type_id,
                quantity,
                quantity_remaining,
                is_included,
                is_singleton,
                raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            contract_id,
            record_id,
            type_id,
            quantity,
            quantity_remaining,
            is_included,
            is_singleton,
            json.dumps(item),
        ))
        if cur.rowcount > 0:
            new_items += 1

    conn.commit()
    if new_items > 0:
        print(f"  Contract {contract_id}: {new_items} new items.")

def sync_industry_jobs(conn: sqlite3.Connection, esi: ESIClient) -> None:
    """
    Sync corporation industry jobs (manufacturing, etc.) from ESI.

    We primarily care about:
      - installer_id
      - product_type_id
      - runs
      - cost
      - status

    We'll later use delivered jobs to credit installers in member_flows.
    """
    cur = conn.cursor()
    corp_id = esi.cfg.corporation_id

    page = 1
    new_jobs = 0

    while True:
        path = f"/corporations/{corp_id}/industry/jobs/"
        try:
            data, headers = esi.get(path, params={"page": page, "include_completed": "true"})
        except requests.HTTPError as ex:
            print(f"Failed to sync industry jobs page {page}: {ex}")
            break

        if not data:
            break

        for row in data:
            job_id = row["job_id"]
            cur.execute("SELECT 1 FROM industry_jobs WHERE job_id = ?", (job_id,))
            exists = cur.fetchone()

            if exists:
                continue

            cur.execute("""
                INSERT INTO industry_jobs (
                    job_id,
                    installer_id,
                    facility_id,
                    activity_id,
                    blueprint_id,
                    blueprint_type_id,
                    product_type_id,
                    runs,
                    cost,
                    status,
                    start_date,
                    end_date,
                    pause_date,
                    completed_character_id,
                    completed_date,
                    successful_runs,
                    location_id,
                    raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                row.get("job_id"),
                row.get("installer_id"),
                row.get("facility_id"),
                row.get("activity_id"),
                row.get("blueprint_id"),
                row.get("blueprint_type_id"),
                row.get("product_type_id"),
                row.get("runs"),
                row.get("cost"),
                row.get("status"),
                row.get("start_date"),
                row.get("end_date"),
                row.get("pause_date"),
                row.get("completed_character_id"),
                row.get("completed_date"),
                row.get("successful_runs"),
                row.get("location_id"),
                json.dumps(row),
            ))
            new_jobs += 1

        x_pages = headers.get("X-Pages")
        if x_pages is None or page >= int(x_pages):
            break
        page += 1

    conn.commit()
    print(f"Synced industry jobs: {new_jobs} new jobs.")

def sync_market_orders(conn: sqlite3.Connection, esi: ESIClient) -> None:
    """
    Sync corporation market orders (open + history) from ESI.

    Uses:
      - /corporations/{corp_id}/orders/
      - /corporations/{corp_id}/orders/history/

    We care about:
      - sell orders (is_buy_order = 0)
      - order issuer (issued_by)
      - price, volume_total, volume_remain
    """
    cur = conn.cursor()
    corp_id = esi.cfg.corporation_id

    def _sync_orders(path_suffix: str, is_history: int) -> int:
        page = 1
        new_orders = 0
        while True:
            path = f"/corporations/{corp_id}{path_suffix}"
            try:
                data, headers = esi.get(path, params={"page": page})
            except requests.HTTPError as ex:
                print(f"Failed to sync market orders ({path_suffix}) page {page}: {ex}")
                break

            if not data:
                break

            for row in data:
                order_id = row["order_id"]
                cur.execute("""
                    INSERT OR REPLACE INTO market_orders (
                        order_id,
                        type_id,
                        location_id,
                        volume_total,
                        volume_remain,
                        price,
                        is_buy_order,
                        issued_by,
                        state,
                        issued,
                        last_updated,
                        range,
                        wallet_division,
                        region_id,
                        is_history,
                        raw_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id,
                    row.get("type_id"),
                    row.get("location_id"),
                    row.get("volume_total"),
                    row.get("volume_remain"),
                    row.get("price"),
                    1 if row.get("is_buy_order") else 0,
                    row.get("issued_by"),
                    row.get("state"),
                    row.get("issued"),
                    row.get("last_updated"),
                    row.get("range"),
                    row.get("wallet_division"),
                    row.get("region_id"),
                    is_history,
                    json.dumps(row),
                ))
                new_orders += 1

            x_pages = headers.get("X-Pages")
            if x_pages is None or page >= int(x_pages):
                break
            page += 1

        return new_orders

    new_open = _sync_orders("/orders/", is_history=0)
    new_hist = _sync_orders("/orders/history/", is_history=1)

    conn.commit()
    print(f"Synced market orders: {new_open} open, {new_hist} history rows.")

# -----------------------
# Janice appraisal helper
# -----------------------

def janice_appraise_contract(
    conn: sqlite3.Connection, esi: ESIClient, cfg: ESIConfig, contract_id: int
) -> Optional[float]:
    """
    Build an item list for a contract, send it to Janice, and store
    immediate split total in contracts.janice_immediate_split.
    """
    if not cfg.janice_api_key:
        return None

    cur = conn.cursor()
    cur.execute("""
        SELECT type_id, quantity
        FROM contract_items
        WHERE contract_id = ? AND is_included = 1
    """, (contract_id,))
    rows = cur.fetchall()
    if not rows:
        return None

    # Resolve type names for each type_id
    items_for_janice: List[Tuple[str, int]] = []
    for type_id, qty in rows:
        name = get_type_name(conn, esi, type_id)
        items_for_janice.append((name, int(qty or 0)))

    # Build item list lines "QTY Name"
    body_lines = [f"{qty} {name}" for (name, qty) in items_for_janice]
    body = "\n".join(body_lines)

    payload = {
        "id": contract_id,
        "method": "Appraisal.create",
        "params": {
            "marketId": 2,
            "designation": 100,
            "pricing": 200,
            "pricingVariant": 100,
            "pricePercentage": 1,
            "input": body,
            "comment": "",
            "compactize": True,
        },
    }

    postRequestBody = json.dumps(payload)

    url = cfg.janice_url or "https://janice.e-351.com/api/rpc/v1?m=Appraisal.create"

    try:
        resp = requests.post(
            url,
            headers={"Content-Type": "text/plain"},
            data=postRequestBody.encode("utf-8"),
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        result = data.get("result", {})
        immediate_split = (
            result.get("immediatePrices", {})
                  .get("totalSplitPrice")
        )

        # Add delay to avoid hammering Janice
        time.sleep(0.75)

        # Store in DB
        cur.execute("""
            UPDATE contracts
            SET janice_immediate_split = ?
            WHERE contract_id = ?
        """, (immediate_split, contract_id))
        conn.commit()
        return immediate_split
    except Exception as ex:
        print(f"  [Janice] Appraisal failed for contract {contract_id}: {ex}")
        return None


# -----------------------
# Member flow helpers
# -----------------------
def get_recent_member_flows(
    conn: sqlite3.Connection,
    limit: int = 500,
    # direction: Optional[str] = None,
    # source: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Fetch recent member_flow rows, optionally filtered by direction/source.
    Returns a list of dicts for easier formatting.
    """
    cur = conn.cursor()

    where_clauses = []
    params: List[Any] = []

    # if direction:
    #     where_clauses.append("direction = ?")
    #     params.append(direction)

    # if source:
    #     where_clauses.append("source = ?")
    #     params.append(source)

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    sql = f"""
        SELECT id,
               character_id,
               direction,
               source,
               value_isk,
               created_at,
               note
        FROM member_flows
        {where_sql}
        ORDER BY datetime(created_at) DESC, id DESC
        LIMIT ?
    """
    params.append(limit)

    cur.execute(sql, params)
    rows = cur.fetchall()

    result: List[Dict[str, Any]] = []
    for rid, char_id, direction, source, value_isk, created_at, note in rows:
        result.append({
            "id": rid,
            "character_id": char_id,
            "direction": direction,
            "source": source,
            "value_isk": float(value_isk or 0.0),
            "created_at": created_at,
            "note": note or "",
        })
    return result

def rebuild_member_flows(conn: sqlite3.Connection) -> None:
    """
    Rebuild member_flows from donations and contracts.

    Direction + source conventions:
      - Wallet donation (player -> corp): direction='in', source='wallet'
      - Member -> Corp contract donation: direction='in', source='contract_in'
      - Corp -> Member subsidy/payout: direction='out', source='contract_out'
    """
    cur = conn.cursor()

    print("Rebuilding member_flows from donations and contracts...")
    cur.execute("DELETE FROM member_flows")

    # Wallet donations -> flows (value in)
    cur.execute("""
        SELECT character_id, journal_esi_id, amount, description
        FROM donations
    """)
    rows = cur.fetchall()
    count_wallet = 0
    for char_id, journal_id, amount, desc in rows:
        if char_id is None or amount is None:
            continue
        cur.execute("""
            INSERT INTO member_flows (
                character_id, direction, source, journal_esi_id, value_isk, note
            )
            VALUES (?, 'in', 'wallet', ?, ?, ?)
        """, (int(char_id), journal_id, float(amount), (desc or "")[:200]))
        count_wallet += 1

    # Member -> Corp donation contracts (value in)
    # for_corporation=1, price=0, finished, with Janice valuation
    cur.execute("""
        SELECT contract_id, issuer_id, janice_immediate_split
        FROM contracts
        WHERE (price IS NULL OR price = 0)
          AND status = 'finished'
          AND janice_immediate_split IS NOT NULL
    """)
    rows = cur.fetchall()
    count_contract_in = 0
    for contract_id, issuer_id, val in rows:
        if issuer_id is None or val is None:
            continue
        cur.execute("""
            INSERT INTO member_flows (
                character_id, direction, source, contract_id, value_isk, note
            )
            VALUES (?, 'in', 'contract_in', ?, ?, 'Donation contract to corp')
        """, (int(issuer_id), contract_id, float(val)))
        count_contract_in += 1

    # Corp -> Member subsidies (value out)
    # for_corporation=1, price>0, finished, Janice valuation exists
    # subsidy = janice_immediate_split - price (if positive)
    cur.execute("""
        SELECT contract_id, assignee_id, janice_immediate_split, price
        FROM contracts
        WHERE for_corporation = 1
          AND price IS NOT NULL
          AND status = 'finished'
          AND janice_immediate_split IS NOT NULL
    """)
    rows = cur.fetchall()
    count_contract_out = 0
    for contract_id, assignee_id, val, price in rows:
        if assignee_id is None or val is None:
            continue
        subsidy = float(val) - float(price or 0)
        if subsidy <= 0:
            continue
        cur.execute("""
            INSERT INTO member_flows (
                character_id, direction, source, contract_id, value_isk, note
            )
            VALUES (?, 'out', 'contract_out', ?, ?, 'Corp subsidy / payout')
        """, (int(assignee_id), contract_id, subsidy))

        cur.execute("""
            INSERT INTO member_flows (
                character_id, direction, source, contract_id, value_isk, note
            )
            VALUES (?, 'in', 'contract_out_subsidy', ?, ?, 'Corp Discount Credit')
        """, (int(assignee_id), contract_id, subsidy * float(.10)))

        count_contract_out += 1

        # Industry manufacturing jobs (value in for installers)
    # Treat job 'cost' as contribution value – proxy for the MFG effort/value.
    cur.execute("""
        SELECT job_id,
               installer_id,
               product_type_id,
               runs,
               cost,
               status
        FROM industry_jobs
        WHERE status = 'delivered'
          AND cost IS NOT NULL
          AND installer_id IS NOT NULL
    """)
    rows = cur.fetchall()
    count_industry = 0
    for job_id, installer_id, product_type_id, runs, cost, status in rows:
        note = f"Industry job {job_id}, type {product_type_id}, runs {runs}"
        cur.execute("""
            INSERT INTO member_flows (
                character_id, direction, source, contract_id, journal_esi_id,
                value_isk, note
            )
            VALUES (?, 'in', 'industry', NULL, NULL, ?, ?)
        """, (int(installer_id), float(cost), note[:200]))
        count_industry += 1

        # Market order sales (1% of realized sell value credited to issuer)
    # Use history orders so we don't double-count partial fills over time.
    cur.execute("""
        SELECT order_id,
               issued_by,
               price,
               volume_total,
               volume_remain,
               is_buy_order,
               state
        FROM market_orders
        WHERE is_history = 1
          AND is_buy_order = 0        -- sell orders only
          AND issued_by IS NOT NULL
          AND price IS NOT NULL
    """)
    rows = cur.fetchall()
    count_market = 0
    for order_id, issued_by, price, volume_total, volume_remain, is_buy_order, state in rows:
        if volume_total is None:
            continue
        sold_volume = float(volume_total) - float(volume_remain or 0)
        if sold_volume <= 0:
            continue

        sale_value = sold_volume * float(price)
        credit = sale_value * 0.01  # 1% market-order credit

        note = f"Market sell order {order_id}, state {state}, sold {sold_volume} @ {price}"
        cur.execute("""
            INSERT INTO member_flows (
                character_id, direction, source, contract_id, journal_esi_id,
                value_isk, note
            )
            VALUES (?, 'in', 'market', NULL, NULL, ?, ?)
        """, (int(issued_by), credit, note[:200]))
        count_market += 1

    conn.commit()
    print(f"  Wallet flows:            {count_wallet}")
    print(f"  Contract donation flows: {count_contract_in}")
    print(f"  Contract subsidy flows:  {count_contract_out}")
    print(f"  Industry job flows:       {count_industry}")
    print(f"  Market order flows:       {count_market}")



def get_flow_totals(conn: sqlite3.Connection) -> Tuple[float, float, float]:
    """
    Returns (total_in, total_out, net) from member_flows.
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN direction = 'in'  THEN value_isk ELSE 0 END), 0),
            COALESCE(SUM(CASE WHEN direction = 'out' THEN value_isk ELSE 0 END), 0)
        FROM member_flows
    """)
    total_in, total_out = cur.fetchone()
    total_in = float(total_in or 0.0)
    total_out = float(total_out or 0.0)
    net = total_in - total_out
    return total_in, total_out, net


def get_member_net_values(conn: sqlite3.Connection) -> Dict[int, float]:
    """
    Build dict: character_id -> net value (in - out) from member_flows.
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT character_id,
               SUM(CASE WHEN direction = 'in' THEN value_isk
                        WHEN direction = 'out' THEN -value_isk
                        ELSE 0 END) AS net_value
        FROM member_flows
        GROUP BY character_id
    """)
    result: Dict[int, float] = {}
    for char_id, net in cur.fetchall():
        if char_id is None:
            continue
        result[int(char_id)] = float(net or 0.0)
    return result


# -----------------------
# CLI commands
# -----------------------
def cmd_report_flows(args, cfg: ESIConfig):
    """
    Print a Discord-friendly recent activity report from member_flows.

    Example line:
    [2025-11-22 17:05][IN][wallet]  Vortakin  +123,456,789 ISK  Some note
    """
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)

    # direction = args.direction
    # source = args.source
    # limit = args.limit

    #flows = get_recent_member_flows(conn, limit=limit, direction=direction, source=source)
    flows = get_recent_member_flows(conn)
    if not flows:
        print("No member flows found.")
        return

    print(f"Recent member flows (limit) — direction={'ANY'}, source={'ANY'}")
    for f in flows:
        char_id = f["character_id"]
        name = get_character_name(conn, esi, char_id) if char_id is not None else "Unknown"

        created = f["created_at"] or ""
        # Trim timestamp to "YYYY-MM-DD HH:MM"
        if created and "T" in created:
            created_disp = created.replace("T", " ")[:16]
        else:
            created_disp = created[:16]

        dir_flag = f["direction"].upper() if f["direction"] else "?"
        src_flag = f["source"] or "?"
        value = f["value_isk"]

        sign = "+" if dir_flag == "IN" else "-"
        value_str = f"{sign}{abs(value):,.2f} ISK"

        # Keep notes short-ish for Discord
        note = f["note"]
        if len(note) > 80:
            note = note[:77] + "..."

        # Format line
        line = f"[{created_disp}][{dir_flag}][{src_flag}] {name:<16} {value_str:>18}  {note}"
        SayItInDiscord(line,cfg)
        print(line)

def cmd_sync_wallet(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)
    sync_wallet_journal(conn, esi)
    update_donations_from_journal(conn)

def cmd_sync_market(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)
    sync_market_orders(conn, esi)

def cmd_list_donations(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    cur = conn.cursor()
    cur.execute("""
        SELECT d.id, d.journal_esi_id, d.character_id, d.amount, d.description, d.processed
        FROM donations d
        ORDER BY d.id DESC
        LIMIT ?
    """, (args.limit,))
    rows = cur.fetchall()
    esi = ESIClient(cfg)

    print(f"{'ID':>4} {'Name':>12}  {'CharID':>12} {'Amount(ISK)':>15} {'Processed':>10}  Description")
    print("-" * 80)
    for row in rows:
        
        id_, esi_id, char_id, amount, desc, processed = row
        name = get_character_name(conn, esi, char_id)
        print(f"{id_:>4} {name:>12} {char_id:>12} {amount:>15,.2f} {bool(processed)!s:>10} {desc[:40]}")
        #print(f"{id_:>4} {char_id:>12} {amount:>15,.2f} {bool(processed)!s:>10}  {desc[:60]}")


def cmd_sync_contracts(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)
    sync_contracts(conn, esi)


def cmd_list_contracts(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)
    cur = conn.cursor()

    cur.execute("""
        SELECT contract_id, type, status, title, date_issued, price, reward, janice_immediate_split
        FROM contracts
        ORDER BY date_issued DESC
        LIMIT ?
    """, (args.limit,))
    contracts = cur.fetchall()

    for cid, ctype, status, title, date_issued, price, reward, imm_split in contracts:
        print("=" * 100)
        print(f"Contract {cid} | Type: {ctype or ''} | Status: {status or ''}")
        print(f"  Title: {title or ''}")
        print(f"  Issued: {date_issued or ''}")
        print(f"  Price:  {price or 0:,.2f} ISK   Reward: {reward or 0:,.2f} ISK")

        # Items for this contract
        cur.execute("""
            SELECT type_id, quantity, quantity_remaining, is_included, is_singleton
            FROM contract_items
            WHERE contract_id = ?
        """, (cid,))
        items = cur.fetchall()

        if not items:
            print("  Items: (none cached)")
        else:
            print("  Items:")
            for type_id, qty, qty_rem, is_included, is_singleton in items:
                name = get_type_name(conn, esi, type_id)
                incl_flag = "" if is_included else " (not included)"
                sing_flag = " [singleton]" if is_singleton else ""
                print(f"    - {qty}x {name}{sing_flag}{incl_flag}")

        # Janice valuation
        if cfg.janice_api_key:
            if imm_split is None:
                imm_split = janice_appraise_contract(conn, esi, cfg, cid)

            if imm_split:
                print(f"Contract {cid} | Type: {ctype or ''} | Status: {status or ''} | Title: {title or ''} | Price:  {price or 0:,.2f} ISK   Reward: {reward or 0:,.2f} ISK for Immediate Split:  {imm_split:,.2f} ISK")
                
                print("  Janice appraisal:")
                print(f"    Immediate Split:  {imm_split:,.2f} ISK")
            else:
                print("  Janice appraisal: (no data / failed)")
        else:
            print("  Janice appraisal: (API key not configured)")

    if not contracts:
        print("No contracts found.")


def cmd_sync_flows(args, cfg: ESIConfig):
    """
    Rebuild member_flows from current donations and contracts.
    """
    conn = init_db(cfg.db_path)
    rebuild_member_flows(conn)


def cmd_dashboard(args, cfg: ESIConfig):
    """
    High-level summary: value in/out and implied shares (net).
    """
    conn = init_db(cfg.db_path)

    total_in, total_out, net = get_flow_totals(conn)
    total_shares = net / SHARE_UNIT_ISK if SHARE_UNIT_ISK > 0 else 0.0

    print("=== Corp Value Flow Dashboard ===")
    print(f"  Total value in:           {total_in:>15,.2f} ISK")
    print(f"  Total value out:          {total_out:>15,.2f} ISK")
    print(f"  -------------------------------{'-' * 10}")
    print(f"  Net value (in - out):     {net:>15,.2f} ISK")
    print(f"  Share unit:               {SHARE_UNIT_ISK:,.0f} ISK per share")
    print(f"  Implied total shares:     {total_shares:,.4f}")
    print()


# def cmd_export_dataset(args, cfg: ESIConfig):
#     """
#     Export per-character net contribution dataset as CSV:
#       character_id,net_value_isk,estimated_shares
#     """
#     conn = init_db(cfg.db_path)
#     per_char = get_member_net_values(conn)

#     print("character_id,net_value_isk,estimated_shares")
#     for char_id, value in sorted(per_char.items(), key=lambda x: x[1], reverse=True):
#         shares = value / SHARE_UNIT_ISK if SHARE_UNIT_ISK > 0 else 0.0
#         print(f"{char_id},{value:.2f},{shares:.6f}")
def cmd_export_dataset(args, cfg: ESIConfig):
    """
    Export per-character net contribution dataset as CSV:
      character_id,character_name,net_value_isk,estimated_shares
    """
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)

    per_char = get_member_net_values(conn)

    print("character_id,character_name,net_value_isk,estimated_shares")
    for char_id, value in sorted(per_char.items(), key=lambda x: x[1], reverse=True):
        name = get_character_name(conn, esi, char_id)
        shares = value / SHARE_UNIT_ISK if SHARE_UNIT_ISK > 0 else 0.0
        # Escape commas in name by wrapping in quotes if needed
        safe_name = f"\"{name}\"" if "," in name else name
        print(f"{char_id},{safe_name},{value:.2f},{shares:.6f}")

def cmd_export_excel(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    output = args.output
    export_all_to_excel(conn, output)

def cmd_sync_industry(args, cfg: ESIConfig):
    conn = init_db(cfg.db_path)
    esi = ESIClient(cfg)
    sync_industry_jobs(conn, esi)

# -----------------------
# Config loading
# -----------------------

def load_config(path: str) -> ESIConfig:
    with open(path, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)

    jan_raw = raw.get("janice", {}) or {}
    custom_raw = raw.get("custom", {}) or {}

    return ESIConfig(
        client_id=raw["esi"]["client_id"],
        client_secret=raw["esi"]["client_secret"],
        refresh_token=raw["esi"]["refresh_token"],
        base_url=raw["esi"].get("base_url", "https://esi.evetech.net/latest"),
        token_url=raw["esi"].get("token_url", "https://login.eveonline.com/v2/oauth/token"),
        corporation_id=int(raw["corp"]["corporation_id"]),
        wallet_division=int(raw["corp"]["wallet_division"]),
        db_path=raw.get("db", {}).get("path", "./ledger.db"),
        janice_api_key=jan_raw.get("api_key"),
        janice_url=jan_raw.get("url", "https://janice.e-351.com/api/rest/v1/pricer"),

        # Add any new custom config values here:
        custom_discord_webhook=custom_raw.get("discord_webhook"),
        custom_share_unit_isk=custom_raw.get("share_unit_isk", 1_000_000_000),
        custom_alert_threshold=custom_raw.get("alert_threshold", 50_000_000),
    )


def main():
    parser = argparse.ArgumentParser(
        description="Ain't Misbehaving Holding Corp Ledger (ESI + Janice)"
    )
    parser.add_argument(
        "--config",
        default="config.yaml",
        help="Path to config.yaml",
    )

    sub = parser.add_subparsers(dest="command", required=True)

    # wallet
    p_sync = sub.add_parser("sync-wallet", help="Sync corp wallet journal and derive donations")
    p_sync.set_defaults(func=cmd_sync_wallet)

    p_list = sub.add_parser("list-donations", help="List recent donations seen via wallet journal")
    p_list.add_argument("--limit", type=int, default=20)
    p_list.set_defaults(func=cmd_list_donations)

    # contracts
    p_csync = sub.add_parser("sync-contracts", help="Sync corporation contracts + items")
    p_csync.set_defaults(func=cmd_sync_contracts)

    p_clist = sub.add_parser(
        "list-contracts",
        help="List recent contracts with items and Janice appraisal"
    )
    p_clist.add_argument("--limit", type=int, default=20)
    p_clist.set_defaults(func=cmd_list_contracts)

    # flows
    p_flows = sub.add_parser(
        "sync-flows",
        help="Rebuild member flows from donations and contracts"
    )
    p_flows.set_defaults(func=cmd_sync_flows)

    # dashboard
    p_dash = sub.add_parser(
        "dashboard",
        help="Show value-in/out summary and implied shares (net)"
    )
    p_dash.set_defaults(func=cmd_dashboard)

    # dataset export
    p_data = sub.add_parser(
        "export-dataset",
        help="Export per-character net value dataset (CSV)"
    )
    p_data.set_defaults(func=cmd_export_dataset)

    p_xl = sub.add_parser("export-excel", help="Export all tables into a multi-sheet Excel workbook")
    p_xl.add_argument(
        "--output",
        default="corp_full_export.xlsx",
        help="Output Excel file path",
    )
    p_xl.set_defaults(func=cmd_export_excel)

    # industry jobs
    p_ind = sub.add_parser(
        "sync-industry",
        help="Sync corporation industry jobs (manufacturing)"
    )
    p_ind.set_defaults(func=cmd_sync_industry)
    
    # market orders
    p_mkt = sub.add_parser(
        "sync-market",
        help="Sync corporation market orders (open + history)"
    )
    p_mkt.set_defaults(func=cmd_sync_market)

    # report orders
    p_rpt = sub.add_parser(
        "report-flows",
        help="Show recent member flows in a Discord-friendly text format"
    )
    p_rpt.set_defaults(func=cmd_report_flows)
    # # report: recent flows, Discord-style
    # p_rpt = sub.add_parser(
    #     "report-flows",
    #     help="Show recent member flows in a Discord-friendly text format"
    # )
    # p_rpt.add_argument(
    #     "--limit",
    #     type=int,
    #     default=1000,
    #     help="Number of recent flows to show (default 1000)",
    # )
    # p_rpt.add_argument(
    #     "--direction",
    #     choices=["in", "out"],
    #     help="Filter by flow direction (in/out)",
    # )
    # p_rpt.add_argument(
    #     "--source",
    #     help="Filter by source (wallet, contract_in, contract_out, industry, market, etc.)",
    # )
    # p_rpt.set_defaults(func=cmd_report_flows)

    args = parser.parse_args()
    cfg = load_config(args.config)
    args.func(args, cfg)


if __name__ == "__main__":
    main()
